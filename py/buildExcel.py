#-*- coding: utf-8 -*-
# 오류처리 및 코드를 최대한 최적화 할 것.
import sys
sys.path.append(r'C:\Python27\Lib')
sys.path.append(r'C:\Python27\Lib\site-packages')

import pymongo
from openpyxl import load_workbook
import json
import os

'''
콘솔에 값을 찍어주는 함수
'''
def stdout(a):
    sys.stdout.write(str(a))

'''
json 파일을 읽어서 객체를 넘겨주는 함수,
여기서 Exception 또는 Validation 처리도 진행해야 됨.
'''
def get_json_obj(document_name):

    file_path = ''.join([os.getcwd(),'\\config\\', document_name, '_XLS.json'])

    with open(file_path) as jsonObj:
        json_data = json.load(jsonObj)

    return json_data

'''
Company json 파일을 읽어서 객체를 넘겨주는 함수,
여기서 Exception 또는 Validation 처리도 진행해야 됨.
'''
def get_company_obj():

    file_path = ''.join([os.getcwd(),'\\config\\Company.json'])

    with open(file_path) as jsonObj:
        json_data = json.load(jsonObj)

    return json_data


'''
Excel 파일을 만드는 함수,
openpyxl 모듈을 이용해서, template 을 읽은 후에 값을 넣고
결과 path 를 보내준다.
'''
def build(filename,document_name,output_path,document_key):
    print filename, '::', document_name, '::', output_path, '::', document_key
    try:
        #Excel 정보를 가지고 있는 Config 파일 읽어오기.
        json_ = get_json_obj(document_name)
        #Company 정보를 가지고 있는 Config 파일 읽어오기.
        cjson_ = get_company_obj()

        #DB 연결에 대한 에러처리 Exception 처리
        collection = db_connection(document_name.lower())
        #Mongo 에서 데이터를 가져오는 구문.
        result = collection.find_one({"VATKEY":document_key})

        if result == None:
            print "결과 값이 없습니다.> [%s]" % document_key
            return 1
        else:
            # 엑셀을 읽어온다.
            wb = load_workbook(filename)
            # 워크시트를 가져온다. (기본은 첫번째 시트를..)
            ws = wb.active

            #Mongo 에서 가져온 값을 넣어준다.
            for items in result.keys():

                if items == '_id' or items == '__v':
                    continue

                if items in json_:
                    ws[json_[items]] = result[items]

            #기본 Company 정보를 넣어준다.
            for citems in cjson_.keys():

                if citems in json_:
                    ws[json_[citems]] = cjson_[citems]

            #Output Path 에 결과 엑셀을 저장한다.
            opath = ''.join([output_path, document_key, '.xlsx'])

            wb.save(opath)

    except OSError:
        pass
    except:
        print "의도되지 않은 오류입니다:",sys.exc_info()[0]
        return 1

    return 0


def db_connection(document_name):
    try:
        client = pymongo.MongoClient()
        db = client.test
        collection = db[document_name]

        return collection
    except pymongo.errors.ConnectionFailure, e:
        return 1


# 메인 함수.
if __name__ == "__main__":
    filename = sys.argv[1]
    output_path = sys.argv[2]
    document_name = sys.argv[3]
    document_key = sys.argv[4]
    result = build(filename,document_name,output_path,document_key)
    stdout(result)
    sys.exit(result)
